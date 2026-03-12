from __future__ import annotations

import base64
from io import BytesIO

from document_engine.model_types import TextBlock


class ExcelExtractor:
    def decode_base64_excel(self, encoded: str) -> bytes:
        try:
            return base64.b64decode(encoded, validate=True)
        except Exception as exc:
            raise ValueError("Invalid base64 Excel payload") from exc

    def extract(
        self,
        excel_bytes: bytes,
        header_axis: str = "first_row",
    ) -> tuple[str, list[TextBlock], int]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError("openpyxl is required for Excel analysis") from exc

        try:
            workbook = load_workbook(
                filename=BytesIO(excel_bytes),
                data_only=True,
                read_only=True,
            )
        except Exception as exc:
            raise ValueError("Invalid or unsupported Excel file. Use native .xlsx/.xlsm") from exc

        blocks: list[TextBlock] = []
        text_parts: list[str] = []
        sheet_count = 0

        axis = header_axis if header_axis in {"first_row", "first_column"} else "first_row"

        for sheet_idx, sheet in enumerate(workbook.worksheets, start=1):
            sheet_count = sheet_idx
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            raw_blocks: list[TextBlock] = []
            raw_text_parts: list[str] = []
            for row_idx, row in enumerate(rows, start=1):
                row_values: list[str] = []
                for col_idx, value in enumerate(row, start=1):
                    cell_text = self._cell_to_text(value)
                    if not cell_text:
                        continue

                    row_values.append(cell_text)
                    x0 = float((col_idx - 1) * 100.0)
                    x1 = x0 + 90.0
                    y_top = float(100000.0 - (row_idx - 1) * 20.0)
                    y_bottom = y_top - 16.0

                    raw_blocks.append(
                        TextBlock(
                            text=cell_text,
                            page=sheet_idx,
                            x0=x0,
                            y0=y_bottom,
                            x1=x1,
                            y1=y_top,
                        )
                    )

                if row_values:
                    raw_text_parts.append(" ".join(row_values))

            structured_pairs = self._build_structured_pairs(rows, axis)
            structured_blocks = self._build_structured_pair_blocks(rows, sheet_idx, axis)

            # Give priority to structured key/value pairs so header axis choice
            # effectively drives matching before generic cell-by-cell detection.
            if structured_pairs:
                text_parts.extend(structured_pairs)
                text_parts.extend(raw_text_parts)
            else:
                text_parts.extend(raw_text_parts)

            blocks.extend(structured_blocks + raw_blocks)

        workbook.close()
        full_text = "\n".join(text_parts)
        return full_text, blocks, max(sheet_count, 1)

    def _cell_to_text(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return " ".join(str(value).split())

    def _build_structured_pairs(self, rows: list[list[object]], axis: str) -> list[str]:
        pairs: list[str] = []
        if not rows:
            return pairs

        max_cols = max((len(r) for r in rows), default=0)
        if axis == "first_column":
            for col_idx in range(1, max_cols):
                for row_idx, row in enumerate(rows, start=1):
                    label = self._cell_to_text(row[0] if row else None)
                    value = self._cell_to_text(row[col_idx] if col_idx < len(row) else None)
                    if label and value:
                        pairs.append(f"{label}: {value}")
        else:
            headers = rows[0] if rows else []
            for row_idx in range(1, len(rows)):
                row = rows[row_idx]
                for col_idx in range(max_cols):
                    label = self._cell_to_text(headers[col_idx] if col_idx < len(headers) else None)
                    value = self._cell_to_text(row[col_idx] if col_idx < len(row) else None)
                    if label and value:
                        pairs.append(f"{label}: {value}")
        return pairs

    def _build_structured_pair_blocks(
        self,
        rows: list[list[object]],
        page: int,
        axis: str,
    ) -> list[TextBlock]:
        blocks: list[TextBlock] = []
        if not rows:
            return blocks

        max_cols = max((len(r) for r in rows), default=0)
        if axis == "first_column":
            for col_idx in range(1, max_cols):
                for row_idx, row in enumerate(rows, start=1):
                    label = self._cell_to_text(row[0] if row else None)
                    value = self._cell_to_text(row[col_idx] if col_idx < len(row) else None)
                    if not label or not value:
                        continue
                    x0 = float((col_idx - 1) * 100.0)
                    x1 = x0 + 90.0
                    y_top = float(100000.0 - (row_idx - 1) * 20.0)
                    y_bottom = y_top - 16.0
                    blocks.append(
                        TextBlock(
                            text=f"{label}: {value}",
                            page=page,
                            x0=x0,
                            y0=y_bottom,
                            x1=x1,
                            y1=y_top,
                        )
                    )
        else:
            headers = rows[0] if rows else []
            for row_idx in range(2, len(rows) + 1):
                row = rows[row_idx - 1]
                for col_idx in range(1, max_cols + 1):
                    label = self._cell_to_text(headers[col_idx - 1] if col_idx - 1 < len(headers) else None)
                    value = self._cell_to_text(row[col_idx - 1] if col_idx - 1 < len(row) else None)
                    if not label or not value:
                        continue
                    x0 = float((col_idx - 1) * 100.0)
                    x1 = x0 + 90.0
                    y_top = float(100000.0 - (row_idx - 1) * 20.0)
                    y_bottom = y_top - 16.0
                    blocks.append(
                        TextBlock(
                            text=f"{label}: {value}",
                            page=page,
                            x0=x0,
                            y0=y_bottom,
                            x1=x1,
                            y1=y_top,
                        )
                    )
        return blocks
