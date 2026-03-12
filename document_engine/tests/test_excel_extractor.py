from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from document_engine.core.excel_extractor import ExcelExtractor


def test_excel_extractor_reads_cells_with_spatial_order() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "KBIS"
    ws["A1"] = "Gestion, direction,"
    ws["A3"] = "nom"
    ws["B3"] = "DUPONT"

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()

    full_text, blocks, page_count = ExcelExtractor().extract(buffer.getvalue())

    assert page_count == 1
    assert "Gestion, direction," in full_text
    assert "nom DUPONT" in full_text

    by_text = {b.text: b for b in blocks}
    assert by_text["nom"].y1 < by_text["Gestion, direction,"].y1
    assert by_text["DUPONT"].x0 > by_text["nom"].x1


def test_excel_extractor_supports_first_column_headers() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "KBIS_COL"
    ws["A1"] = "champ"
    ws["B1"] = "valeur"
    ws["A2"] = "president_nom"
    ws["B2"] = "DURAND"

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()

    full_text, _, _ = ExcelExtractor().extract(buffer.getvalue(), header_axis="first_column")
    assert "president_nom: DURAND" in full_text


def test_excel_header_axis_changes_structured_output() -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "col_a"
    ws["B1"] = "col_b"
    ws["A2"] = "row_2_a"
    ws["B2"] = "row_2_b"

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    excel_bytes = buffer.getvalue()

    by_row_text, _, _ = ExcelExtractor().extract(excel_bytes, header_axis="first_row")
    by_col_text, _, _ = ExcelExtractor().extract(excel_bytes, header_axis="first_column")

    assert "col_a: row_2_a" in by_row_text
    assert "row_2_a: row_2_b" in by_col_text
    assert "col_a: row_2_a" not in by_col_text
