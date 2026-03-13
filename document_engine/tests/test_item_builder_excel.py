from __future__ import annotations

import base64
from io import BytesIO

from openpyxl import Workbook

from document_engine.training.item_builder import ItemBuilder


def _excel_b64_with_pairs() -> str:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Nom"
    ws["B1"] = "Jean Dupont"
    ws["A2"] = "Prénom"
    ws["B2"] = "Jean"
    ws["A3"] = "Numéro de dossier"
    ws["B3"] = 3

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return base64.b64encode(buffer.getvalue()).decode("ascii")


def test_item_builder_excel_uses_labels_for_required_elements() -> None:
    doc = _excel_b64_with_pairs()
    config = ItemBuilder().build(
        item="kbis_excel",
        template="contract",
        language="fr",
        threshold=0.7,
        docs=[doc, doc, doc],
        document_type="excel",
        excel_header_axis="first_column",
    )

    names = [entry["name"] for entry in config["required_elements"]]
    assert "Nom" in names
    assert "Prénom" in names
    assert "Numéro de dossier" in names
    assert config["audit"]["document_type"] == "excel"
    assert config["audit"]["excel_pairs_preview"]
